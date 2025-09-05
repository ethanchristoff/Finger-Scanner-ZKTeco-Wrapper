import io
from typing import Any
from datetime import datetime, time, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from adms_wrapper.core.data_processing import process_attendance_summary
from adms_wrapper.core.db_queries import (
    get_device_branch_mappings,
    get_employee_branch_mappings,
    get_employee_designation_mappings,
    get_employee_name_mappings,
    get_shift_templates,
    get_default_shift,
    get_setting,
)

NOON_HOUR = 12


def map_branch(sn: Any, mappings_df: pd.DataFrame) -> str:
    """Map device serial number to branch name."""
    if pd.isna(sn) or sn is None:
        return ""

    # If mappings_df is None or empty, nothing to map
    if mappings_df is None or mappings_df.empty:
        return ""

    # Work on a copy to avoid mutating caller DataFrame
    df = mappings_df.copy()

    # Ensure expected columns exist; attempt common alternatives
    if "serial_number" not in df.columns or "branch_name" not in df.columns:
        rename_map = {}
        for alt in ("serial", "sn", "device_sn", "serial_no"):
            if alt in df.columns and "serial_number" not in df.columns:
                rename_map[alt] = "serial_number"
                break
        for alt in ("branch", "branchname", "branch_name_str"):
            if alt in df.columns and "branch_name" not in df.columns:
                rename_map[alt] = "branch_name"
                break
        if rename_map:
            df = df.rename(columns=rename_map)

    if "serial_number" not in df.columns or "branch_name" not in df.columns:
        return ""

    # Compare as strings to avoid dtype mismatch
    try:
        sn_str = str(sn).strip()
        match = df[df["serial_number"].astype(str).str.strip() == sn_str]
    except Exception:
        try:
            match = df[df["serial_number"].astype(str).str.contains(str(sn), na=False)]
        except Exception:
            return ""

    if not match.empty:
        return match.iloc[0]["branch_name"]
    return ""


def map_designation(emp_id: Any, designation_df: pd.DataFrame) -> str:
    """Map employee ID to designation."""
    if pd.isna(emp_id) or emp_id is None or designation_df.empty:
        return ""
    if "employee_id" not in designation_df.columns:
        return ""
    row = designation_df[designation_df["employee_id"] == str(emp_id)]
    if not row.empty:
        return row.iloc[0]["designation"]
    return ""


def map_employee_branch(emp_id: Any, employee_branch_df: pd.DataFrame) -> str:
    """Map employee ID to branch name."""
    if pd.isna(emp_id) or emp_id is None or employee_branch_df.empty:
        return ""
    if "employee_id" not in employee_branch_df.columns:
        return ""
    row = employee_branch_df[employee_branch_df["employee_id"] == str(emp_id)]
    if not row.empty:
        return row.iloc[0]["branch_name"]
    return ""
    return ""


def map_employee_name(emp_id: Any, employee_name_df: pd.DataFrame) -> str:
    """Map employee ID to employee name."""
    if pd.isna(emp_id) or emp_id is None or employee_name_df.empty:
        return ""
    if "employee_id" not in employee_name_df.columns:
        return ""
    row = employee_name_df[employee_name_df["employee_id"] == str(emp_id)]
    if not row.empty:
        return row.iloc[0]["employee_name"]
    return ""


def _to_time(obj: Any) -> time | None:
    """Normalize various time/datetime/string inputs to a time object (HH:MM)."""
    if pd.isna(obj) or obj is None or str(obj) == "":
        return None
    try:
        if hasattr(obj, "time") and hasattr(obj, "date"):
            return obj.time()
        # If it's a pandas Timestamp
        if hasattr(obj, "hour") and hasattr(obj, "minute") and not hasattr(obj, "date"):
            return time(int(obj.hour), int(obj.minute))
        s = str(obj)
        # Accept formats like '08:30:00' or '08:30'
        s = s.strip()
        if " " in s:
            s = s.split(" ")[1]
        parts = s.split(":")
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        return time(hour, minute)
    except Exception:
        return None


def determine_shift_flag(start_time: Any, end_time: Any, shift_start: Any, shift_end: Any) -> str:
    """Determine shift flag based on times.

    Statuses returned: 'normal', 'late in', 'early out', 'overtime'.
    Rules:
    - In is 'normal' if at or before shift_start or within 5 minutes after shift_start; otherwise 'late in'.
    - Out is 'early out' if before shift_end.
    - Out is 'normal' if between shift_end and shift_end + 15 minutes (inclusive).
    - Out is 'overtime' if after shift_end + 15 minutes.
    """
    flag = "normal"
    try:
        s_time = _to_time(start_time)
        e_time = _to_time(end_time)
        sh_start = _to_time(shift_start)
        sh_end = _to_time(shift_end)

        today = datetime.today().date()

        # Build shift start/end datetimes. If shift crosses midnight (end <= start) treat end as next day.
        sh_start_dt = datetime.combine(today, sh_start) if sh_start else None
        sh_end_dt = datetime.combine(today, sh_end) if sh_end else None
        if sh_start_dt and sh_end_dt and sh_end_dt <= sh_start_dt:
            sh_end_dt = sh_end_dt + timedelta(days=1)

        # Helper to convert a time to a datetime in the same logical window as the shift
        def _to_dt_with_shift(t: time | None) -> datetime | None:
            if not t or not sh_start_dt:
                return datetime.combine(today, t) if t else None
            dt = datetime.combine(today, t)
            # If shift crosses midnight and the time is earlier than shift start, assume it's on the next day
            if sh_end_dt and sh_end_dt.date() != sh_start_dt.date() and dt < sh_start_dt:
                dt = dt + timedelta(days=1)
            return dt

        # Determine late in using datetimes (robust across midnight)
        # Note: We no longer flag early check-ins as special - they are normal
        if s_time and sh_start:
            try:
                s_dt = _to_dt_with_shift(s_time)
                # Only check for late arrivals (after grace period)
                late_threshold = sh_start_dt + timedelta(minutes=5)
                if s_dt and s_dt > late_threshold:
                    # Checked in more than 5 minutes after shift start -> late in
                    flag = "late in"
            except Exception:
                # best-effort fallback to time-only compare
                try:
                    late_in_threshold = (datetime.combine(today, sh_start) + timedelta(minutes=5)).time()
                    if s_time > late_in_threshold:
                        flag = "late in"
                except Exception:
                    # leave flag as-is
                    pass

        # Determine out-related flags using datetimes for correctness across midnight
        if e_time and sh_end:
            try:
                e_dt = _to_dt_with_shift(e_time)
                # Early out: any checkout before the official shift end or before the shift start
                # (e.g., someone checked out very early on the same day)
                if (sh_start_dt and e_dt < sh_start_dt) or e_dt < sh_end_dt:
                    flag = "early out"
                else:
                    # Use configurable late checkout grace from settings (minutes after shift end to consider normal)
                    try:
                        grace_minutes = int(get_setting("late_checkout_grace_minutes") or 15)
                    except Exception:
                        grace_minutes = 15
                    normal_threshold_dt = sh_end_dt + timedelta(minutes=grace_minutes)
                    # Any checkout after the normal grace window is considered a late checkout
                    if e_dt <= normal_threshold_dt:
                        # Within normal grace window
                        if flag != "late in":
                            flag = "normal"
                    else:
                        # After normal grace window -> late checkout
                        flag = "late checkout"
            except Exception:
                # Fallback to original time-only logic: consider early out if checkout is before shift start or before shift end
                try:
                    if sh_start and e_time < sh_start:
                        flag = "early out"
                    elif e_time < sh_end:
                        flag = "early out"
                    else:
                        try:
                            grace_minutes = int(get_setting("late_checkout_grace_minutes") or 15)
                        except Exception:
                            grace_minutes = 15
                        normal_out_threshold = (datetime.combine(today, sh_end) + timedelta(minutes=grace_minutes)).time()
                        try:
                            if e_time <= normal_out_threshold:
                                if flag != "late in":
                                    flag = "normal"
                            else:
                                flag = "late checkout"
                        except Exception:
                            flag = "overtime"
                except Exception:
                    # If we cannot compare to sh_start, fall back to previous behavior
                    if e_time < sh_end:
                        flag = "early out"
                    else:
                        # Use configurable late checkout grace from settings (minutes after shift end to consider normal)
                        try:
                            grace_minutes = int(get_setting("late_checkout_grace_minutes") or 15)
                        except Exception:
                            grace_minutes = 15
                        normal_out_threshold = (datetime.combine(today, sh_end) + timedelta(minutes=grace_minutes)).time()
                        try:
                            if e_time <= normal_out_threshold:
                                if flag != "late in":
                                    flag = "normal"
                            else:
                                flag = "overtime"
                        except Exception:
                            flag = "overtime"
                else:
                    try:
                        grace_minutes = int(get_setting("late_checkout_grace_minutes") or 15)
                    except Exception:
                        grace_minutes = 15
                    normal_out_threshold = (datetime.combine(today, sh_end) + timedelta(minutes=grace_minutes)).time()
                    try:
                        if e_time <= normal_out_threshold:
                            if flag != "late in":
                                flag = "normal"
                        else:
                            flag = "late checkout"
                    except Exception:
                        flag = "overtime"
    except Exception:
        pass
    return flag


def determine_no_shift_flag(end_time: Any) -> str:
    """Determine shift flag for employees with no shift assignment.

    This function is kept for backward compatibility, but the main code now prefers using
    a configurable default shift from settings (if present) or falling back to 08:00-17:30.
    """
    flag = "normal"
    try:
        if pd.notna(end_time) and str(end_time) != "":
            end_time_obj = _to_time(end_time)
            if end_time_obj:
                # If checkout is in the afternoon, consider overtime; otherwise early out
                if end_time_obj >= time(12, 0):
                    flag = "overtime"
                else:
                    flag = "early out"
    except Exception:
        pass
    return flag


def get_shift_info_with_capped(emp_id: str, work_status: str, start_time: Any, end_time: Any, shift_df: pd.DataFrame) -> tuple[str, str]:
    """Get shift information for an employee without shift_capped parameter."""
    # Determine shift template to use: prefer assigned shift in shift_df; else use configured default
    chosen_shift_name = ""
    chosen_shift_start = None
    chosen_shift_end = None

    # Try to find assigned shift
    if not shift_df.empty:
        shift_row = shift_df[shift_df["user_id"] == str(emp_id)]
        if not shift_row.empty:
            chosen_shift_name = shift_row.iloc[0]["shift_name"]
            chosen_shift_start = shift_row.iloc[0]["shift_start"]
            chosen_shift_end = shift_row.iloc[0]["shift_end"]

    # If no assigned shift, try configured default shift template
    if not chosen_shift_start or not chosen_shift_end:
        default_shift_name = get_default_shift() or ""
        if default_shift_name:
            # Look up template by name
            try:
                templates = get_shift_templates() or []
                for t in templates:
                    if t.get("shift_name") == default_shift_name:
                        chosen_shift_name = t.get("shift_name")
                        chosen_shift_start = t.get("shift_start")
                        chosen_shift_end = t.get("shift_end")
                        break
            except Exception:
                pass

    # If still missing, fall back to 08:00 - 17:30 default
    if not chosen_shift_start or not chosen_shift_end:
        chosen_shift_name = chosen_shift_name or "Default"
        chosen_shift_start = chosen_shift_start or time(8, 0)
        chosen_shift_end = chosen_shift_end or time(17, 30)

    if work_status == "absent":
        return chosen_shift_name, "absent"

    # If there's no end_time but there is a start_time, check for shift cap (no checkout within configured hours after shift end)
    if (pd.isna(end_time) or end_time is None or str(end_time) == "") and pd.notna(start_time):
        try:
            if hasattr(start_time, "date"):
                date_part = start_time.date()
            else:
                date_part = datetime.today().date()

            sh_end = _to_time(chosen_shift_end)
            if sh_end:
                shift_end_dt = datetime.combine(date_part, sh_end)
                try:
                    # Get the grace period for late checkout
                    grace_minutes = int(get_setting("late_checkout_grace_minutes") or 15)
                except Exception:
                    grace_minutes = 15

                try:
                    cap_hours = int(get_setting("shift_cap_hours") or 8)
                except Exception:
                    cap_hours = 8

                # Add grace minutes to shift end before calculating cap time
                grace_adjusted_end = shift_end_dt + timedelta(minutes=grace_minutes)
                # Then add cap hours
                cap_dt = grace_adjusted_end + timedelta(hours=cap_hours)
                if datetime.now() >= cap_dt:
                    return chosen_shift_name, "no checkout"
        except Exception:
            pass

    # If an end_time exists, check whether it's within the 8-hour window; determine_shift_flag will classify early/normal/overtime/shift_capped
    flag = determine_shift_flag(start_time, end_time, chosen_shift_start, chosen_shift_end)
    return chosen_shift_name, flag


def apply_branch_mappings(summary_df: pd.DataFrame) -> pd.DataFrame:
    """Apply device branch mappings to summary DataFrame."""
    mappings = get_device_branch_mappings() or []
    mappings_df = pd.DataFrame(mappings)
    # If no mappings available, set empty columns
    if mappings_df is None or mappings_df.empty:
        summary_df["start_device_sn_branch"] = ""
        summary_df["end_device_sn_branch"] = ""
        return summary_df

    summary_df["start_device_sn_branch"] = summary_df.apply(
        lambda row: map_branch(row.get("start_device_sn"), mappings_df) if row.get("work_status") == "worked" else "",
        axis=1,
    )
    summary_df["end_device_sn_branch"] = summary_df.apply(
        lambda row: map_branch(row.get("end_device_sn"), mappings_df) if row.get("work_status") == "worked" else "",
        axis=1,
    )
    return summary_df


def apply_designation_mappings(summary_df: pd.DataFrame) -> pd.DataFrame:
    """Apply designation mappings to summary DataFrame."""
    designation_mappings = get_employee_designation_mappings() or []
    designation_df = pd.DataFrame(designation_mappings)
    if not summary_df.empty and "employee_id" in summary_df.columns:
        summary_df["designation"] = summary_df["employee_id"].apply(lambda emp_id: map_designation(emp_id, designation_df))
    else:
        summary_df["designation"] = ""
    return summary_df


def apply_employee_branch_mappings(summary_df: pd.DataFrame) -> pd.DataFrame:
    """Apply employee branch mappings to summary DataFrame."""
    employee_branch_mappings = get_employee_branch_mappings() or []
    employee_branch_df = pd.DataFrame(employee_branch_mappings)
    if not summary_df.empty and "employee_id" in summary_df.columns:
        summary_df["employee_branch"] = summary_df["employee_id"].apply(lambda emp_id: map_employee_branch(emp_id, employee_branch_df))
    else:
        summary_df["employee_branch"] = ""
    return summary_df


def apply_employee_name_mappings(summary_df: pd.DataFrame) -> pd.DataFrame:
    """Apply employee name mappings to summary DataFrame."""
    employee_name_mappings = get_employee_name_mappings() or []
    employee_name_df = pd.DataFrame(employee_name_mappings)
    if not summary_df.empty and "employee_id" in summary_df.columns:
        summary_df["employee_name"] = summary_df["employee_id"].apply(lambda emp_id: map_employee_name(emp_id, employee_name_df))
    else:
        summary_df["employee_name"] = ""
    return summary_df


def apply_shift_mappings(summary_df: pd.DataFrame, shift_mappings: list[dict[str, Any]] | None) -> pd.DataFrame:
    """Apply shift mappings to summary DataFrame."""
    shift_df = pd.DataFrame(shift_mappings) if shift_mappings else pd.DataFrame()

    summary_df["shift_name"] = ""
    summary_df["shift_flag"] = ""
    # New columns to track early/late check-ins separately from shift_flag
    summary_df["late_in"] = False
    summary_df["late_in_time"] = ""

    for idx, row in summary_df.iterrows():
        no_checkout = row.get("no_checkout", False)

        if no_checkout:
            shift_name, _ = get_shift_info_with_capped(row["employee_id"], row["work_status"], row["start_time"], row["end_time"], shift_df)
            flag = "no checkout"
        else:
            shift_name, flag = get_shift_info_with_capped(row["employee_id"], row["work_status"], row["start_time"], row["end_time"], shift_df)

        # If early checkout is true, it takes priority over any non "no checkout" flag
        try:
            if not no_checkout and bool(row.get("early_checkout", False)):
                flag = "early out"
        except Exception:
            pass

        summary_df.loc[idx, "shift_name"] = shift_name
        summary_df.loc[idx, "shift_flag"] = flag

        try:
            s_time = _to_time(row.get("start_time"))
            sh_start = _to_time(
                shift_name and shift_df[shift_df["user_id"] == str(row["employee_id"])].iloc[0]["shift_start"]
                if not shift_df.empty and not shift_df[shift_df["user_id"] == str(row["employee_id"])].empty
                else shift_name and None
            )
        except Exception:
            s_time = None
            sh_start = None

        if s_time is None or sh_start is None:
            try:
                # Try to find shift template by name
                if shift_name:
                    templates = get_shift_templates() or []
                    for t in templates:
                        if t.get("shift_name") == shift_name:
                            sh_start = _to_time(t.get("shift_start"))
                            break
            except Exception:
                pass

        late_in_flag = False
        if s_time and sh_start:
            try:
                today = datetime.today().date()
                s_dt = datetime.combine(today, s_time)
                late_threshold_dt = datetime.combine(today, sh_start) + timedelta(minutes=5)
                if s_dt > late_threshold_dt:
                    late_in_flag = True
            except Exception:
                late_in_flag = False

        summary_df.loc[idx, "late_in"] = late_in_flag
        if late_in_flag:
            try:
                summary_df.loc[idx, "late_in_time"] = s_time.strftime("%H:%M:%S") if s_time else str(row.get("start_time") or "")
            except Exception:
                summary_df.loc[idx, "late_in_time"] = str(row.get("start_time") or "")

    return summary_df


def clean_attendance_summary(summary_df: pd.DataFrame) -> pd.DataFrame:
    """Clean attendance summary by removing unwanted columns and formatting times."""
    if summary_df.empty:
        return summary_df

    # Format start_time and end_time to show only time (not date)
    if "start_time" in summary_df.columns:
        summary_df["start_time"] = summary_df["start_time"].apply(lambda x: x.strftime("%H:%M:%S") if pd.notna(x) and hasattr(x, "strftime") else "")

    if "end_time" in summary_df.columns:
        summary_df["end_time"] = summary_df["end_time"].apply(lambda x: x.strftime("%H:%M:%S") if pd.notna(x) and hasattr(x, "strftime") else "")

    # Remove unwanted columns (keeping start_device_sn_branch and end_device_sn_branch)
    columns_to_remove = ["start_device_sn", "end_device_sn", "no_checkout", "designation", "employee_branch"]

    for col in columns_to_remove:
        if col in summary_df.columns:
            summary_df = summary_df.drop(columns=[col])

    return summary_df


def create_subtotal_rows(summary_df: pd.DataFrame) -> list[dict[str, Any]]:
    """Create subtotal rows for each employee with separate days worked and total hours columns."""
    output_rows = []

    for emp_id, group in summary_df.groupby("employee_id", sort=False):
        output_rows.extend(group.to_dict(orient="records"))

        worked_group = group[group["work_status"] == "worked"].copy()
        days_worked = len(worked_group)

        if not worked_group.empty:
            worked_group.loc[:, "time_spent_td"] = pd.to_timedelta(worked_group["time_spent"])
            subtotal = worked_group["time_spent_td"].sum()
            subtotal_str = str(subtotal).split(".")[0]
        else:
            subtotal_str = "0:00:00"

        subtotal_row = dict.fromkeys(summary_df.columns, "")
        subtotal_row["employee_id"] = emp_id
        if not group.empty:
            subtotal_row["employee_name"] = group.iloc[0].get("employee_name", "")
        subtotal_row["day"] = "Subtotal"
        subtotal_row["days_worked"] = days_worked
        subtotal_row["total_hours"] = subtotal_str
        subtotal_row["work_status"] = "subtotal"
        output_rows.append(subtotal_row)

    return output_rows


def filter_out_sundays_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """Filter out Sunday entries from a DataFrame."""
    if df.empty or "day" not in df.columns:
        return df

    # Create a mask to identify non-Sunday rows
    mask = df["day"].apply(lambda day: True if day == "Subtotal" else pd.to_datetime(day).weekday() != 6)

    return df[mask]


def generate_attendance_summary(
    attendences: list[dict[str, Any]],
    _device_logs: list[dict[str, Any]],
    _finger_logs: list[dict[str, Any]],
    _migration_logs: list[dict[str, Any]],
    _user_logs: list[dict[str, Any]],
    shift_mappings: list[dict[str, Any]] | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> pd.DataFrame:
    """Generate attendance summary with all mappings applied, excluding Sundays."""
    summary_df = process_attendance_summary(attendences, start_date, end_date)

    if summary_df is not None and not summary_df.empty:
        # Filter out Sundays before applying mappings
        summary_df = filter_out_sundays_from_df(summary_df)

        summary_df = apply_branch_mappings(summary_df)
        summary_df = apply_designation_mappings(summary_df)
        summary_df = apply_employee_branch_mappings(summary_df)
        summary_df = apply_employee_name_mappings(summary_df)
        summary_df = apply_shift_mappings(summary_df, shift_mappings)

        # Clean the summary before creating subtotal rows
        summary_df = clean_attendance_summary(summary_df)

        output_rows = create_subtotal_rows(summary_df)
        merged = pd.DataFrame(output_rows, columns=[*summary_df.columns.tolist(), "days_worked", "total_hours"])
    else:
        merged = summary_df if summary_df is not None else pd.DataFrame()

    return merged


def find_column_indices(ws: Any) -> tuple[int | None, int | None, int | None]:
    """Find column indices for highlighting."""
    day_col = None
    work_status_col = None
    no_checkout_col = None

    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        header = str(cell.value).strip().lower()
        # Accept both 'day' and common export header 'date'
        if header in ("day", "date"):
            day_col = idx
        if header in ("work_status", "work status", "workstatus"):
            work_status_col = idx
        if header in ("no_checkout", "no checkout", "shift capped", "shiftcap"):
            no_checkout_col = idx

    return day_col, work_status_col, no_checkout_col


def apply_subtotal_highlighting(row: list, day_col: int | None, blue_fill: PatternFill) -> None:
    """Apply highlighting for subtotal rows."""
    if not day_col:
        return
    try:
        cell_val = row[day_col - 1].value
        if cell_val is None:
            return
        if str(cell_val).strip().lower() == "subtotal":
            for cell in row:
                cell.fill = blue_fill
    except Exception:
        return


def apply_status_highlighting(row: list, work_status_col: int | None, no_checkout_col: int | None, green_fill: PatternFill, red_fill: PatternFill, yellow_fill: PatternFill) -> None:
    """Apply highlighting based on work status."""
    if not work_status_col:
        return

    work_status = row[work_status_col - 1].value
    no_checkout = row[no_checkout_col - 1].value if no_checkout_col else False

    if no_checkout:
        for cell in row:
            cell.fill = yellow_fill
    elif work_status == "worked":
        for cell in row:
            cell.fill = green_fill
    elif work_status == "absent":
        for cell in row:
            cell.fill = red_fill


def apply_row_highlighting(ws: Any) -> None:
    """Apply highlighting to rows based on work status."""
    day_col, work_status_col, no_checkout_col = find_column_indices(ws)

    blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        apply_subtotal_highlighting(row, day_col, blue_fill)
        apply_status_highlighting(row, work_status_col, no_checkout_col, green_fill, red_fill, yellow_fill)


def apply_flag_highlighting(ws: Any) -> None:
    """Apply highlighting based on the Shift Flag column for Excel-only exports.

    Colors applied (per-row):
    - overtime -> light orange
    - late in -> light red/pink
    - shift_capped -> light purple
    - normal/other -> light green
    """
    # Find the column index for 'Shift Flag' (case-insensitive) and 'Shift Flag' alternatives
    flag_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).strip().lower() in ("shift flag", "shift_flag", "shiftflag"):
            flag_col_idx = idx
            break

    if not flag_col_idx:
        return

    # Define fills
    overtime_fill = PatternFill(start_color="FFDFA6", end_color="FFDFA6", fill_type="solid")  # light orange
    late_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red/pink
    shift_capped_fill = PatternFill(start_color="E6E0FF", end_color="E6E0FF", fill_type="solid")  # light purple
    normal_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        try:
            cell = row[flag_col_idx - 1]
            val_raw = cell.value or ""
            val = str(val_raw).strip().lower()

            # Determine presence of keywords (overlapping flags allowed)
            has_overtime = "overtime" in val or "over time" in val
            has_late_in = "late in" in val or "latein" in val
            has_no_checkout = "no checkout" in val or "shift cap" in val or "shiftcap" in val

            # Choose a fill with precedence: overtime > no_checkout > late_in > normal
            if has_overtime:
                fill = overtime_fill
            elif has_no_checkout:
                fill = shift_capped_fill
            elif has_late_in:
                fill = late_fill
            else:
                fill = normal_fill

            for c in row:
                c.fill = fill
        except Exception:
            # If anything fails for a row, skip coloring that row
            continue


def create_employee_summary_sheet(summary_df: pd.DataFrame) -> pd.DataFrame:
    """Create a comprehensive summary sheet with employee statistics including days worked, subtotal hours, etc."""
    if summary_df.empty:
        return pd.DataFrame()

    summary_rows = []

    for emp_id, group in summary_df.groupby("employee_id", sort=False):
        # Get employee info from the first row of the group
        first_row = group.iloc[0]

        # Calculate statistics
        total_days = len(group)
        worked_days = len(group[group["work_status"] == "worked"])
        absent_days = len(group[group["work_status"] == "absent"])

        # Calculate total worked hours
        worked_group = group[group["work_status"] == "worked"].copy()
        if not worked_group.empty:
            worked_group.loc[:, "time_spent_td"] = pd.to_timedelta(worked_group["time_spent"])
            subtotal = worked_group["time_spent_td"].sum()
            subtotal_str = str(subtotal).split(".")[0]

            # Calculate average hours per worked day
            avg_hours_per_day = subtotal / worked_days if worked_days > 0 else pd.Timedelta(0)
            avg_hours_str = str(avg_hours_per_day).split(".")[0]
        else:
            subtotal_str = "0:00:00"
            avg_hours_str = "0:00:00"

        # Get shift information
        shift_name = first_row.get("shift_name", "")

        # Count different shift flags (only on worked days), normalizing names and using booleans when available
        worked_mask = group["work_status"].astype(str).str.lower() == "worked"

        # Normalize shift_flag values for robust counting across pipelines
        flag_series = group.get("shift_flag", pd.Series([""] * len(group), index=group.index))
        flag_norm = (
            flag_series.astype(str)
            .str.strip()
            .str.lower()
            .replace(
                {
                    "shift capped": "no checkout",
                    "shift cap": "no checkout",
                    "shiftcap": "no checkout",
                    "latein": "late in",
                    "earlyin": "early in",
                    "earlyout": "early out",
                    "over time": "overtime",
                }
            )
        )

        # Prefer boolean late_in when present; otherwise fall back to flag
        late_in_count = int(group.loc[worked_mask, "late_in"].fillna(False).astype(bool).sum()) if "late_in" in group.columns else int((flag_norm.eq("late in") & worked_mask).sum())

        early_out_count = int((flag_norm.eq("early out") & worked_mask).sum())

        # Treat both 'late checkout' and 'overtime' as late checkout days in the summary
        late_checkout_count = int((flag_norm.isin(["late checkout", "overtime"]) & worked_mask).sum())

        # Treat both 'on time' and 'normal' as on-time days
        on_time_count = int((flag_norm.isin(["on time", "normal"]) & worked_mask).sum())

        # Safely count rows marked as no_checkout. If the column is missing, treat as zero.
        if "no_checkout" in group.columns:
            try:
                no_checkout_count = int(group.loc[worked_mask, "no_checkout"].fillna(False).astype(bool).sum())
            except Exception:
                # If conversion fails for any unexpected reason, fallback to len of truthy values
                no_checkout_count = int(sum(1 for v in group.loc[worked_mask, "no_checkout"] if bool(v)))
        else:
            # Fallback: count rows where normalized flag indicates 'no checkout'
            no_checkout_count = int((flag_norm.eq("no checkout") & worked_mask).sum())

        # Calculate attendance percentage
        attendance_percentage = (worked_days / total_days * 100) if total_days > 0 else 0

        # Create summary row
        summary_row = {
            "employee_id": emp_id,
            "employee_name": first_row.get("employee_name", ""),
            "shift_name": shift_name,
            "total_days": total_days,
            "days_worked": worked_days,
            "days_absent": absent_days,
            "attendance_percentage": f"{attendance_percentage:.1f}%",
            "total_hours_worked": subtotal_str,
            "avg_hours_per_day": avg_hours_str,
            "on_time_days": on_time_count,
            "late_in_days": late_in_count,
            "early_out_days": early_out_count,
            "late_checkout_days": late_checkout_count,
            "no_checkout_days": no_checkout_count,
        }
        summary_rows.append(summary_row)

    # Create DataFrame and sort by employee_id
    summary_summary_df = pd.DataFrame(summary_rows)
    if not summary_summary_df.empty:
        # Convert employee_id to string to avoid type mismatch during sorting
        summary_summary_df["employee_id"] = summary_summary_df["employee_id"].astype(str)
        summary_summary_df = summary_summary_df.sort_values("employee_id")

    return summary_summary_df


def write_excel(
    attendences: list[dict[str, Any]], device_logs: list[dict[str, Any]], finger_logs: list[dict[str, Any]], migration_logs: list[dict[str, Any]], user_logs: list[dict[str, Any]], merged: pd.DataFrame
) -> io.BytesIO:
    """Write data to Excel file with formatting."""
    output = io.BytesIO()

    # Generate employee summary sheet
    employee_summary = create_employee_summary_sheet(merged)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(attendences).to_excel(writer, sheet_name="Attendences", index=False)
        pd.DataFrame(device_logs).to_excel(writer, sheet_name="DeviceLogs", index=False)
        pd.DataFrame(finger_logs).to_excel(writer, sheet_name="FingerLogs", index=False)
        pd.DataFrame(migration_logs).to_excel(writer, sheet_name="Migrations", index=False)
        pd.DataFrame(user_logs).to_excel(writer, sheet_name="Users", index=False)
        # Create an export-only DataFrame using the requested column order and names
        # Do not mutate `merged` (the in-memory DataFrame used by the UI)
        if merged is None:
            export_df = pd.DataFrame(
                columns=[
                    "EPF Number",
                    "Employee Name",
                    "In Time",
                    "Out Time",
                    "Working Hours",
                    "Work Status",
                    "In Location",
                    "Out Location",
                    "Shift Flag",
                    "Total Work Dates",
                    "Total Work Hours",
                ]
            )
        else:
            # Helper to safely get a column series or a blank series when missing
            def _col_series(df: pd.DataFrame, col_name: str):
                if col_name in df.columns:
                    return df[col_name]
                # Try some common alternative column names
                alt_map = {
                    "employee_id": ["EPF Number", "EPF No", "employee_id"],
                    "employee_name": ["Employee Name", "employee_name"],
                    "start_time": ["start_time", "In Time", "in_time"],
                    "end_time": ["end_time", "Out Time", "out_time"],
                    "time_spent": ["time_spent", "Working Hours", "working_hours"],
                    "work_status": ["work_status", "Work Status"],
                    "start_device_sn_branch": ["start_device_sn_branch", "In Location"],
                    "end_device_sn_branch": ["end_device_sn_branch", "Out Location"],
                    "shift_flag": ["shift_flag", "Shift Flag"],
                    "days_worked": ["days_worked", "Total Work Dates"],
                    "total_hours": ["total_hours", "Total Work Hours"],
                }
                # If df has any of the mapped alternatives, return it
                for alt in alt_map.get(col_name, []):
                    if alt in df.columns:
                        return df[alt]
                # Otherwise return blank series matching the length
                return pd.Series([""] * len(df), index=df.index)

            export_df = pd.DataFrame(
                {
                    "EPF Number": _col_series(merged, "employee_id"),
                    "Employee Name": _col_series(merged, "employee_name"),
                    "Date": _col_series(merged, "day"),
                    "In Time": _col_series(merged, "start_time"),
                    "Out Time": _col_series(merged, "end_time"),
                    "Working Hours": _col_series(merged, "time_spent"),
                    "Work Status": _col_series(merged, "work_status"),
                    "In Location": _col_series(merged, "start_device_sn_branch"),
                    "Out Location": _col_series(merged, "end_device_sn_branch"),
                    # We'll compute a combined Shift Flag that includes late-in if present
                    "Shift Flag": _col_series(merged, "shift_flag"),
                    "Total Work Dates": _col_series(merged, "days_worked"),
                    "Total Work Hours": _col_series(merged, "total_hours"),
                }
            )

            # Combine shift_flag with late_in so overlapping flags are visible
            try:
                # ensure late_in series exists
                late_series = merged["late_in"].fillna(False).astype(bool) if "late_in" in merged.columns else pd.Series([False] * len(merged), index=merged.index)

                # get current shift flag column (may be named differently)
                base_flags = export_df["Shift Flag"].fillna("").astype(str) if "Shift Flag" in export_df.columns else pd.Series([""] * len(export_df), index=export_df.index)

                combined_flags = []
                for idx, bf in base_flags.items():
                    parts = []
                    bf_str = str(bf).strip()
                    if bf_str:
                        parts.extend([p.strip() for p in bf_str.split(";") if p.strip()])
                    # only append 'late in' if not already present
                    if late_series.iloc[idx] and not any(p.lower() == "late in" for p in parts):
                        parts.append("late in")

                    combined = "; ".join(parts) if parts else ""
                    combined_flags.append(combined)

                export_df["Shift Flag"] = pd.Series(combined_flags, index=export_df.index)
            except Exception:
                # If combining fails, leave the original Shift Flag column as-is
                pass

        export_df.to_excel(writer, sheet_name="AttendanceSummary", index=False)
        employee_summary.to_excel(writer, sheet_name="EmployeeSummary", index=False)

    output.seek(0)
    wb = load_workbook(output)

    if "AttendanceSummary" in wb.sheetnames:
        ws = wb["AttendanceSummary"]
        # Apply our excel-specific flag-based highlighting (overtime, late in, no checkout)
        try:
            apply_flag_highlighting(ws)
        except Exception:
            # If highlighting fails, continue without breaking export
            pass
        # Also apply existing row highlighting for other markers (subtotals / work_status)
        apply_row_highlighting(ws)

        # Configure print/page setup to better fit an A4 page for printing
        try:
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        except Exception:
            # openpyxl may not support some attributes on older versions; ignore if it fails
            pass

    new_output = io.BytesIO()
    wb.save(new_output)
    new_output.seek(0)
    return new_output
