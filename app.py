from flask import Flask, render_template, request, send_file, redirect, url_for, flash
import pandas as pd
from adms_wrapper.core.db_queries import get_attendences, get_device_logs, get_finger_log, get_migrations, get_users, get_device_branch_mappings, add_device_branch_mapping, delete_device_branch_mapping
from adms_wrapper.__main__ import process_attendance_summary

import io
import os
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Needed for flash messages
@app.route("/device_branch_mapping", methods=["GET", "POST"])
def device_branch_mapping():
    if request.method == "POST":
        # Handle deletion
        delete_sn = request.form.get("delete_serial")
        if delete_sn:
            delete_device_branch_mapping(delete_sn)
            flash(f"Mapping deleted: {delete_sn}", "success")
            return redirect(url_for("device_branch_mapping"))
        # Handle addition
        serial_number = request.form.get("serial_number")
        branch_name = request.form.get("branch_name")
        if serial_number and branch_name:
            add_device_branch_mapping(serial_number, branch_name)
            flash(f"Mapping added: {serial_number} → {branch_name}", "success")
        else:
            flash("Both serial number and branch name are required.", "error")
        return redirect(url_for("device_branch_mapping"))
    mappings = get_device_branch_mappings() or []
    return render_template("device_branch_mapping.html", mappings=mappings)

# Ensure the static and templates folders exist
if not os.path.exists("static"):
    os.makedirs("static")
if not os.path.exists("templates"):
    os.makedirs("templates")

@app.route("/", methods=["GET"])
def index():
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    attendences = get_attendences() or []
    # Filter attendences by date if provided
    if start_date or end_date:
        df = pd.DataFrame(attendences)
        if 'timestamp' in df.columns:
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            if start_date:
                df = df[df['timestamp'] >= pd.to_datetime(start_date)]
            if end_date:
                df = df[df['timestamp'] <= pd.to_datetime(end_date)]
            attendences = df.to_dict(orient="records")
    summary_df = process_attendance_summary(attendences)
    summary = summary_df.to_dict(orient="records") if summary_df is not None else []
    return render_template("dashboard.html", attendences=attendences, summary=summary, start_date=start_date or '', end_date=end_date or '')

@app.route("/download_xlsx")
def download_xlsx():
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    # Use the same logic as the main function to filter and export
    output = io.BytesIO()
    attendences = get_attendences() or []
    if start_date or end_date:
        df = pd.DataFrame(attendences)
        if 'timestamp' in df.columns:
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            if start_date:
                df = df[df['timestamp'] >= pd.to_datetime(start_date)]
            if end_date:
                df = df[df['timestamp'] <= pd.to_datetime(end_date)]
            attendences = df.to_dict(orient="records")
    device_logs = get_device_logs() or []
    finger_logs = get_finger_log() or []
    migration_logs = get_migrations() or []
    user_logs = get_users() or []
    summary_df = process_attendance_summary(attendences)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(attendences).to_excel(writer, sheet_name="Attendences", index=False)
        pd.DataFrame(device_logs).to_excel(writer, sheet_name="DeviceLogs", index=False)
        pd.DataFrame(finger_logs).to_excel(writer, sheet_name="FingerLogs", index=False)
        pd.DataFrame(migration_logs).to_excel(writer, sheet_name="Migrations", index=False)
        pd.DataFrame(user_logs).to_excel(writer, sheet_name="Users", index=False)
        if summary_df is not None:
            # Calculate weekly total work time per user and append as 'Total' row
            df_sum = summary_df.copy()
            # Convert time_spent string to timedelta
            df_sum['time_spent_td'] = pd.to_timedelta(df_sum['time_spent'])
            # Sum durations per employee
            total_df = (
                df_sum.groupby('employee_id')
                .agg({'time_spent_td': 'sum'})
                .reset_index()
            )
            # Format total time and assign 'Total' label for day
            total_df['day'] = 'Total'
            total_df['time_spent'] = total_df['time_spent_td'].apply(lambda x: str(x).split('.')[0])
            # Clear other columns
            for col in ['start_time', 'end_time', 'start_device_sn', 'end_device_sn']:
                total_df[col] = ''
            # Reorder columns to match summary_df
            total_df = total_df[summary_df.columns]
            # Combine original summary with totals
            out_df = pd.concat([summary_df, total_df], ignore_index=True)
            out_df.to_excel(writer, sheet_name="AttendanceSummary", index=False)

        # --- Aggregate Sheet for Unique User Logins per Day ---
        if summary_df is not None and not summary_df.empty:
            agg = summary_df.copy()
            agg['date'] = pd.to_datetime(agg['day'])
            agg['date'] = agg['date'].dt.strftime('%Y-%m-%d')
            # Prepare serial_number from summary and merge branch mappings
            agg['serial_number'] = agg.get('start_device_sn', None)
            mappings = get_device_branch_mappings() or []
            mappings_df = pd.DataFrame(mappings)
            if not mappings_df.empty:
                agg = agg.merge(mappings_df, how='left', on='serial_number')
            # Ensure branch_name column exists and fill missing
            if 'branch_name' not in agg.columns:
                agg['branch_name'] = ''
            else:
                agg['branch_name'] = agg['branch_name'].fillna('')
            agg_sheet = agg.groupby(['employee_id', 'date', 'serial_number', 'branch_name']).agg(
                first_time=('start_time', 'first'),
                last_time=('end_time', 'last')
            ).reset_index()
            def get_status(row):
                if pd.isna(row['first_time']) and pd.isna(row['last_time']):
                    return 'no activity'
                elif pd.isna(row['first_time']):
                    return 'missing check-in'
                elif pd.isna(row['last_time']):
                    return 'missing check-out'
                else:
                    return 'ok'
            agg_sheet['status'] = agg_sheet.apply(get_status, axis=1)
            agg_sheet.to_excel(writer, sheet_name="Aggregate", index=False)
    output.seek(0)
    wb = load_workbook(output)
    # Highlight 'Total' rows in AttendanceSummary sheet in dark blue
    if 'AttendanceSummary' in wb.sheetnames:
        ws_sum = wb['AttendanceSummary']
        # Find the 'day' column index
        day_col = None
        for idx, cell in enumerate(ws_sum[1], start=1):
            if cell.value == 'day':
                day_col = idx
                break
        if day_col:
            blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            # Iterate data rows
            for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row, min_col=1, max_col=ws_sum.max_column):
                if row[day_col-1].value == 'Total':
                    for cell in row:
                        cell.fill = blue_fill
    if 'Aggregate' in wb.sheetnames:
        ws = wb['Aggregate']
        status_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'status':
                status_col = idx
                break
        if status_col:
            green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for row in ws.iter_rows(min_row=2, min_col=status_col, max_col=status_col):
                cell = row[0]
                if cell.value == 'ok':
                    for c in ws[cell.row]:
                        c.fill = green
                elif cell.value == 'missing check-in' or cell.value == 'missing check-out':
                    for c in ws[cell.row]:
                        c.fill = yellow
                elif cell.value == 'no activity':
                    for c in ws[cell.row]:
                        c.fill = red
    new_output = io.BytesIO()
    wb.save(new_output)
    new_output.seek(0)
    return send_file(new_output, as_attachment=True, download_name="output.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)

